"""Parse an uploaded recipients file (CSV / TXT / XLSX) into a clean email list.

We don't care which column emails live in or whether they're mixed with other
text — every cell / line gets regex-scanned, every match deduped and
validated. The caller gets back the clean list plus a breakdown of what we
threw away and why, so users can fix their source data.
"""
import csv
import io
import re
from typing import Dict, List

from auth import validate_email

# Strict email shape — used to PULL likely candidates from arbitrary text.
EMAIL_REGEX = re.compile(
    r"[A-Za-z0-9._%+\-!#$&'*/=?`{|}~^]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}",
    re.UNICODE,
)
# Loose "has an @" shape — anything with @ but no whitespace on either side.
# We use this to catch rows the user MEANT as an email but typoed (missing
# TLD, double @, etc) so we can report them as rejects instead of silently
# skipping them.
LOOSE_AT_REGEX = re.compile(r"\S*@\S*", re.UNICODE)

MAX_FILE_BYTES = 5 * 1024 * 1024  # 5 MB — plenty for ~50k addresses

SUPPORTED_EXTS = {"csv", "txt", "tsv", "xlsx"}


def _detect_kind(filename: str, content_type: str) -> str:
    ext = (filename or "").lower().rsplit(".", 1)[-1] if "." in (filename or "") else ""
    if ext == "xlsx":
        return "xlsx"
    if ext == "xls":
        return "xls_legacy"
    if ext == "csv":
        return "csv"
    if ext in ("txt", "tsv"):
        return "text"
    # MIME fallbacks
    ct = (content_type or "").lower()
    if "spreadsheetml" in ct or ct.endswith("/xlsx"):
        return "xlsx"
    if "csv" in ct:
        return "csv"
    if ct.startswith("text/"):
        return "text"
    return "text"  # last-ditch — try to scan as plain text


def _process_chunk(chunk: str, valid: List[str], invalid: List[Dict],
                     seen_valid: set, seen_invalid: set) -> None:
    """Process one logical 'row' of source content (a text line / a whole CSV
    row / an XLSX row). Mutates valid + invalid in place.

    Rules:
      * Blank or no-@ chunks (headers, notes, separator rows) → silent skip.
      * Strict email match → validate_email → either valid or invalid bucket.
      * @ present but no strict match → reported as invalid with a clear
        reason ('no parseable email — missing TLD?' etc.) so the user can fix.
    """
    if not chunk:
        return
    cleaned = chunk.strip()
    if not cleaned:
        return
    if "@" not in cleaned:
        return  # header row, note, blank, or non-email content — silent skip

    # Strict matches first — those are our real candidates.
    strict = EMAIL_REGEX.findall(cleaned)

    # Anything with @ that DIDN'T strict-match is a probable typo.
    loose_remaining = []
    if strict:
        # Remove strict hits from the cleaned text to find what's left over
        remainder = cleaned
        for s in strict:
            remainder = remainder.replace(s, " ", 1)
        loose_remaining = [
            m for m in LOOSE_AT_REGEX.findall(remainder)
            if "@" in m and m not in strict
        ]
    else:
        loose_remaining = LOOSE_AT_REGEX.findall(cleaned)

    for c in strict:
        norm = c.strip(".,;:()<>[]'\"").lower()
        if not norm:
            continue
        err = validate_email(norm)
        if err:
            if norm not in seen_invalid:
                seen_invalid.add(norm)
                invalid.append({"value": norm, "reason": err})
        elif norm not in seen_valid:
            seen_valid.add(norm)
            valid.append(norm)

    for c in loose_remaining:
        bad = c.strip(".,;:()<>[]'\"").strip()
        if not bad or "@" not in bad:
            continue
        # Diagnose the typo
        if bad.count("@") != 1:
            reason = "multiple or zero @ symbols"
        elif "." not in bad.split("@", 1)[1]:
            reason = "missing TLD (no dot after @)"
        elif bad.startswith("@") or bad.endswith("@"):
            reason = "empty local part or domain"
        else:
            reason = "doesn't match a valid email pattern"
        norm = bad.lower()
        if norm not in seen_invalid:
            seen_invalid.add(norm)
            invalid.append({"value": bad[:80], "reason": reason})


def _walk_text(text: str, valid: List[str], invalid: List[Dict],
                 seen_valid: set, seen_invalid: set) -> None:
    for line in (text or "").splitlines():
        _process_chunk(line, valid, invalid, seen_valid, seen_invalid)


def _walk_csv(content: bytes, valid: List[str], invalid: List[Dict],
                seen_valid: set, seen_invalid: set) -> None:
    text = content.decode("utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    for row in reader:
        # Join the row so a name+email in adjacent columns survives the regex
        joined = " ".join(c or "" for c in row)
        _process_chunk(joined, valid, invalid, seen_valid, seen_invalid)


def _walk_xlsx(content: bytes, valid: List[str], invalid: List[Dict],
                 seen_valid: set, seen_invalid: set) -> None:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            joined = " ".join("" if c is None else str(c) for c in row)
            _process_chunk(joined, valid, invalid, seen_valid, seen_invalid)


def parse_recipients(content: bytes, content_type: str, filename: str) -> Dict:
    """Returns:
        {
          kind: 'csv' | 'xlsx' | 'text' | 'xls_legacy',
          valid: [str, ...],      # clean, deduped, lowercased
          invalid: [{value, reason}, ...],  # rejected candidates, capped at 50
          invalid_truncated: bool,
          invalid_total: int,
          valid_count: int,
        }
    Raises ValueError for unsupported formats / over-size files.
    """
    if not content:
        raise ValueError("Uploaded file is empty.")
    if len(content) > MAX_FILE_BYTES:
        raise ValueError(
            f"File too large ({len(content):,} bytes). "
            f"Max is {MAX_FILE_BYTES // 1024 // 1024} MB."
        )

    kind = _detect_kind(filename, content_type)
    if kind == "xls_legacy":
        raise ValueError(
            "Legacy .xls files aren't supported. Re-save as .xlsx or export to .csv first."
        )

    valid: List[str] = []
    invalid: List[Dict] = []
    seen_valid: set = set()
    seen_invalid: set = set()

    if kind == "xlsx":
        try:
            _walk_xlsx(content, valid, invalid, seen_valid, seen_invalid)
        except Exception as e:
            raise ValueError(f"Couldn't read this .xlsx file ({e}). Try exporting it to .csv.")
    elif kind == "csv":
        _walk_csv(content, valid, invalid, seen_valid, seen_invalid)
    else:
        _walk_text(content.decode("utf-8", errors="ignore"), valid, invalid, seen_valid, seen_invalid)

    return {
        "kind": kind,
        "valid_count": len(valid),
        "valid": valid,
        "invalid_total": len(invalid),
        "invalid": invalid[:50],
        "invalid_truncated": len(invalid) > 50,
    }
